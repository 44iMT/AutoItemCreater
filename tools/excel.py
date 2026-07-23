"""
Excel 工具：读取 / 创建 / 追加 Excel 文件（支持 .xls / .xlsx）
"""
import json
import os
from openpyxl import Workbook, load_workbook


def read_excel(file_path: str, sheet_name: str, max_rows: int) -> str:
    """
    读取 Excel 文件内容并返回制表符分隔的表格文本。支持 .xls 和 .xlsx 格式。

    参数:
        file_path:  Excel 文件路径，比如 'D:/data/商品列表.xlsx'
        sheet_name: 工作表名称，比如 'Sheet1'
        max_rows:   最多读取行数，推荐 20-50，避免一次返回太多数据
    """
    print(f"[excel] 读取 '{file_path}' sheet='{sheet_name}' max={max_rows}")
    if not os.path.exists(file_path):
        return f"文件不存在: {file_path}"

    ext = os.path.splitext(file_path)[1].lower()
    lines = []
    row_count = 0

    # ---- .xls 旧版格式 ----
    if ext == ".xls":
        try:
            import xlrd
        except ImportError:
            return "需要 xlrd 库来读取 .xls 文件，请执行: pip install xlrd"

        try:
            wb = xlrd.open_workbook(file_path)
        except Exception as e:
            return f"打开文件失败: {e}"

        if sheet_name not in wb.sheet_names():
            return f"工作表 '{sheet_name}' 不存在，可用的: {', '.join(wb.sheet_names())}"

        ws = wb.sheet_by_name(sheet_name)
        for r in range(ws.nrows):
            if row_count >= max_rows:
                lines.append("... (还有更多行，已截断)")
                break
            cells = [str(ws.cell_value(r, c)) if ws.cell_value(r, c) != "" else "" for c in range(ws.ncols)]
            lines.append("\t".join(cells))
            row_count += 1

    # ---- .xlsx 新版格式 ----
    else:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            return f"打开文件失败: {e}"

        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"工作表 '{sheet_name}' 不存在，可用的: {', '.join(wb.sheetnames)}"

        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if row_count >= max_rows:
                lines.append("... (还有更多行，已截断)")
                break
            cells = [str(v) if v is not None else "" for v in row]
            lines.append("\t".join(cells))
            row_count += 1
        wb.close()

    result = "\n".join(lines)
    print(f"[excel] 读取完成，{row_count} 行")
    return result


def create_excel(file_path: str, sheet_name: str, headers: str, rows: str) -> str:
    """
    创建新的 Excel 文件并写入数据。

    参数:
        file_path:  Excel 文件路径，比如 'D:/data/导出商品.xlsx'
        sheet_name: 工作表名称，比如 '商品列表'
        headers:    列标题，逗号分隔，比如 '商品名称,条码,价格,规格'
        rows:       数据行，JSON 二维数组格式，每行是一个数组。
                    比如 '[["可乐","123","3.5","500ml"],["雪碧","456","3.5","500ml"]]'
    """
    print(f"[excel] 创建 '{file_path}' sheet='{sheet_name}'")
    header_list = [h.strip() for h in headers.split(",")]

    try:
        data_rows = json.loads(rows)
    except json.JSONDecodeError as e:
        return f"rows 参数 JSON 格式错误: {e}"

    if not isinstance(data_rows, list):
        return "rows 必须是 JSON 数组格式"

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(header_list)
        for row in data_rows:
            if isinstance(row, list):
                ws.append(row)
            else:
                ws.append([str(row)])
        wb.save(file_path)
    except Exception as e:
        return f"创建失败: {e}"

    print(f"[excel] 创建完成，{len(data_rows)} 行数据")
    return f"成功创建 {file_path}，{len(header_list)} 列 {len(data_rows)} 行"


def append_rows(file_path: str, sheet_name: str, rows: str) -> str:
    """
    向已有 Excel 文件追加数据行。

    参数:
        file_path:  Excel 文件路径，比如 'D:/data/商品列表.xlsx'
        sheet_name: 工作表名称
        rows:       数据行，JSON 二维数组格式。
                    比如 '[["可乐","123","3.5"],["雪碧","456","3.5"]]'
    """
    print(f"[excel] 追加 '{file_path}' sheet='{sheet_name}'")
    if not os.path.exists(file_path):
        return f"文件不存在: {file_path}，请先用 create_excel 创建"

    try:
        data_rows = json.loads(rows)
    except json.JSONDecodeError as e:
        return f"rows 参数 JSON 格式错误: {e}"

    if not isinstance(data_rows, list):
        return "rows 必须是 JSON 数组格式"

    try:
        wb = load_workbook(file_path)
    except Exception as e:
        return f"打开文件失败: {e}"

    if sheet_name not in wb.sheetnames:
        wb.close()
        return f"工作表 '{sheet_name}' 不存在，可用的: {', '.join(wb.sheetnames)}"

    ws = wb[sheet_name]
    for row in data_rows:
        if isinstance(row, list):
            ws.append(row)
        else:
            ws.append([str(row)])

    try:
        wb.save(file_path)
    except Exception as e:
        wb.close()
        return f"保存失败: {e}"

    wb.close()
    print(f"[excel] 追加完成，{len(data_rows)} 行")
    return f"成功追加 {len(data_rows)} 行到 {file_path}"
